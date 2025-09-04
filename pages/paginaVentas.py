# pages/ventas.py
import streamlit as st
import couchdb_utils
import os
from datetime import datetime, timezone, timedelta
import base64
import pytz
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fpdf import FPDF
from fpdf.enums import XPos, YPos
import io
try:
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    OpenpyxlImage = None

# --- Configuración Inicial ---
st.set_page_config(layout="wide", page_title="Reportes de Ventas", page_icon="../assets/LOGO.png")
archivo_actual_relativo = os.path.join(os.path.basename(os.path.dirname(__file__)), os.path.basename(__file__))
couchdb_utils.generarLogin(archivo_actual_relativo)

# --- Funciones de Zona Horaria ---
def get_timezone():
    """Retorna la zona horaria de El Salvador"""
    return pytz.timezone('America/El_Salvador')

def convert_to_local_time(utc_datetime_str):
    """Convierte una fecha UTC a hora local de El Salvador"""
    if not utc_datetime_str:
        return None
    try:
        if isinstance(utc_datetime_str, str):
            if utc_datetime_str.endswith('Z'):
                utc_datetime_str = utc_datetime_str.replace('Z', '+00:00')
            utc_dt = datetime.fromisoformat(utc_datetime_str)
        else:
            utc_dt = utc_datetime_str
        
        if utc_dt.tzinfo is None:
            utc_dt = utc_dt.replace(tzinfo=timezone.utc)
        elif utc_dt.tzinfo != timezone.utc:
            utc_dt = utc_dt.astimezone(timezone.utc)
        
        local_tz = get_timezone()
        local_dt = utc_dt.astimezone(local_tz)
        return local_dt
    except Exception:
        return None

# --- Funciones de Reportes ---
def generar_reporte_excel_completo(ordenes_pagadas, ordenes_anuladas, mesas, meseros, tickets, periodo_nombre, start_date=None, end_date=None, resumen_stats=None):
    """Genera un reporte completo en Excel con órdenes pagadas, anuladas y estadísticas"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # --- Hoja de Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Título del resumen
    ws_resumen.cell(row=1, column=1, value=f"RESUMEN DE VENTAS - {periodo_nombre}")
    ws_resumen.cell(row=1, column=1).font = Font(bold=True, size=16)
    
    if start_date and end_date:
        ws_resumen.cell(row=2, column=1, value=f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
    
    # Agregar estadísticas si se proporcionan
    if resumen_stats:
        row = 4
        ws_resumen.cell(row=row, column=1, value="ESTADÍSTICAS GENERALES").font = Font(bold=True)
        row += 1
        
        stats_data = [
            ("📊 Órdenes Pagadas:", resumen_stats.get('ordenes_pagadas', 0)),
            ("💰 Total Ventas Brutas:", f"${resumen_stats.get('ventas_brutas', 0):.2f}"),
            ("🍽️ Items Vendidos:", resumen_stats.get('items_vendidos', 0)),
            ("", ""),
            ("🗑️ Órdenes Anuladas:", resumen_stats.get('ordenes_anuladas', 0)),
            ("💸 Total Perdido:", f"${resumen_stats.get('total_perdido', 0):.2f}"),
            ("🚫 Items Anulados:", resumen_stats.get('items_anulados', 0)),
            ("", ""),
            ("📈 Ventas Netas:", f"${resumen_stats.get('ventas_netas', 0):.2f}"),
        ]
        
        for label, value in stats_data:
            if label:  # Skip empty rows
                ws_resumen.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws_resumen.cell(row=row, column=2, value=value)
            row += 1
    
    # --- Hoja de Productos Vendidos ---
    ws_productos = wb.create_sheet("Productos Vendidos")
    
    # Encabezados productos
    headers_productos = ["Plato Vendido", "Cantidad", "Precio Venta", "Total Venta", "Fecha de Venta", "Mesero", "Método de Pago"]
    for col, header in enumerate(headers_productos, 1):
        cell = ws_productos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Datos productos vendidos
    row = 2
    for orden in ordenes_pagadas:
        fecha_venta = convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion')))
        fecha_str = fecha_venta.strftime('%d/%m/%Y %H:%M') if fecha_venta else 'N/A'
        mesero_nombre = meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')
        
        # Buscar el ticket correspondiente para obtener método de pago
        ticket_orden = next((ticket for ticket in tickets if ticket.get('orden_id') == orden.get('_id')), {})
        pago_info = ticket_orden.get('pago_info', {})
        metodo_pago = pago_info.get('metodo', 'No especificado')
        
        # Formatear método de pago para mejor legibilidad
        if metodo_pago == 'efectivo':
            metodo_display = 'Efectivo'
        elif metodo_pago == 'tarjeta':
            tipo_tarjeta = pago_info.get('tipo_tarjeta', '').title()
            metodo_display = f'Tarjeta ({tipo_tarjeta})'
        elif metodo_pago == 'mixto':
            metodo_display = 'Pago Mixto'
        elif metodo_pago == 'transferencia':
            metodo_display = 'Transferencia'
        elif metodo_pago == 'criptomoneda':
            tipo_cripto = pago_info.get('tipo_cripto', 'Criptomoneda')
            metodo_display = f'Cripto ({tipo_cripto})'
        else:
            metodo_display = metodo_pago.title()
        
        for item in orden.get('items', []):
            # Skip cancelled/anulled items
            if item.get('anulado', False):
                continue
            ws_productos.cell(row=row, column=1, value=item.get('nombre', 'N/A'))
            ws_productos.cell(row=row, column=2, value=item.get('cantidad', 0))
            ws_productos.cell(row=row, column=3, value=item.get('precio_unitario', 0))
            ws_productos.cell(row=row, column=4, value=item.get('cantidad', 0) * item.get('precio_unitario', 0))
            ws_productos.cell(row=row, column=5, value=fecha_str)
            ws_productos.cell(row=row, column=6, value=mesero_nombre)
            ws_productos.cell(row=row, column=7, value=metodo_display)
            row += 1
    
    # --- Hoja de Tickets ---
    ws_tickets = wb.create_sheet("Tickets")
    
    # Encabezados tickets
    headers_tickets = ["Numero Orden", "Numero Ticket", "Mesa", "Mesero", "Total Ticket", "Fecha Venta", "Método de Pago"]
    for col, header in enumerate(headers_tickets, 1):
        cell = ws_tickets.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Datos tickets
    row = 2
    for orden in ordenes_pagadas:
        fecha_venta = convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion')))
        fecha_str = fecha_venta.strftime('%d/%m/%Y %H:%M') if fecha_venta else 'N/A'
        
        # Buscar el ticket correspondiente para obtener método de pago
        ticket_orden = next((ticket for ticket in tickets if ticket.get('orden_id') == orden.get('_id')), {})
        pago_info = ticket_orden.get('pago_info', {})
        metodo_pago = pago_info.get('metodo', 'No especificado')
        
        # Formatear método de pago para mejor legibilidad
        if metodo_pago == 'efectivo':
            metodo_display = 'Efectivo'
        elif metodo_pago == 'tarjeta':
            tipo_tarjeta = pago_info.get('tipo_tarjeta', '').title()
            metodo_display = f'Tarjeta ({tipo_tarjeta})'
        elif metodo_pago == 'mixto':
            metodo_display = 'Pago Mixto'
        elif metodo_pago == 'transferencia':
            metodo_display = 'Transferencia'
        elif metodo_pago == 'criptomoneda':
            tipo_cripto = pago_info.get('tipo_cripto', 'Criptomoneda')
            metodo_display = f'Cripto ({tipo_cripto})'
        else:
            metodo_display = metodo_pago.title()
        
        ws_tickets.cell(row=row, column=1, value=orden.get('numero_orden', 'N/A'))
        ws_tickets.cell(row=row, column=2, value=orden.get('numero_orden', 'N/A'))  # Usando mismo numero como ticket
        ws_tickets.cell(row=row, column=3, value=mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A'))
        ws_tickets.cell(row=row, column=4, value=meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A'))
        ws_tickets.cell(row=row, column=5, value=orden.get('total', 0))
        ws_tickets.cell(row=row, column=6, value=fecha_str)
        ws_tickets.cell(row=row, column=7, value=metodo_display)
        row += 1
    
    # --- Hoja de Órdenes Anuladas ---
    if ordenes_anuladas:
        ws_anuladas = wb.create_sheet("Órdenes Anuladas")
        
        # Encabezados órdenes anuladas
        headers_anuladas = ["Orden #", "Mesa", "Mesero", "Total Perdido", "Fecha Anulación", "Anulada Por", "Motivo", "Productos Anulados"]
        for col, header in enumerate(headers_anuladas, 1):
            cell = ws_anuladas.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="d32f2f", end_color="d32f2f", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Datos órdenes anuladas
        row = 2
        for orden in ordenes_anuladas:
            fecha_anulacion = convert_to_local_time(orden.get('fecha_anulacion_completa'))
            fecha_str = fecha_anulacion.strftime('%d/%m/%Y %H:%M') if fecha_anulacion else 'N/A'
            
            # Calcular total perdido
            items_no_anulados = [item for item in orden.get('items', []) if not item.get('anulado', False)]
            total_perdido = sum(item.get('cantidad', 0) * item.get('precio_unitario', 0) for item in items_no_anulados)
            
            # Lista de productos anulados
            productos_lista = ", ".join([f"{item.get('cantidad', 0)}x {item.get('nombre', 'N/A')}" for item in items_no_anulados])
            
            ws_anuladas.cell(row=row, column=1, value=orden.get('numero_orden', 'N/A'))
            ws_anuladas.cell(row=row, column=2, value=mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A'))
            ws_anuladas.cell(row=row, column=3, value=meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A'))
            ws_anuladas.cell(row=row, column=4, value=total_perdido)
            ws_anuladas.cell(row=row, column=5, value=fecha_str)
            ws_anuladas.cell(row=row, column=6, value=orden.get('usuario_anula_completa', 'N/A'))
            ws_anuladas.cell(row=row, column=7, value=orden.get('motivo_solicitud_anulacion_completa', 'N/A'))
            ws_anuladas.cell(row=row, column=8, value=productos_lista)
            row += 1
    
    # --- Hoja de Anulados (Items anulados individualmente) ---
    ws_anulados = wb.create_sheet("Anulados")
    
    # Encabezados anulados
    headers_anulados = ["Orden #", "Mesa", "Mesero", "Producto Anulado", "Cantidad", "Precio Unit.", "Total Perdido", "Fecha Anulación", "Motivo"]
    for col, header in enumerate(headers_anulados, 1):
        cell = ws_anulados.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ff9800", end_color="ff9800", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Datos anulados (buscar items individuales anulados)
    row = 2
    for orden in ordenes_pagadas:
        items_anulados_individuales = [item for item in orden.get('items', []) if item.get('anulado', False)]
        if items_anulados_individuales:
            for item in items_anulados_individuales:
                fecha_anulacion = convert_to_local_time(orden.get('fecha_creacion'))  # Usar fecha de creación como referencia
                fecha_str = fecha_anulacion.strftime('%d/%m/%Y %H:%M') if fecha_anulacion else 'N/A'
                
                ws_anulados.cell(row=row, column=1, value=orden.get('numero_orden', 'N/A'))
                ws_anulados.cell(row=row, column=2, value=mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A'))
                ws_anulados.cell(row=row, column=3, value=meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A'))
                ws_anulados.cell(row=row, column=4, value=item.get('nombre', 'N/A'))
                ws_anulados.cell(row=row, column=5, value=item.get('cantidad', 0))
                ws_anulados.cell(row=row, column=6, value=item.get('precio_unitario', 0))
                ws_anulados.cell(row=row, column=7, value=item.get('cantidad', 0) * item.get('precio_unitario', 0))
                ws_anulados.cell(row=row, column=8, value=fecha_str)
                ws_anulados.cell(row=row, column=9, value=item.get('motivo_anulacion', 'No especificado'))
                row += 1
    
    # También buscar en órdenes anuladas completas para items individuales anulados antes de la anulación completa
    for orden in ordenes_anuladas:
        items_anulados_individuales = [item for item in orden.get('items', []) if item.get('anulado', False)]
        if items_anulados_individuales:
            for item in items_anulados_individuales:
                fecha_anulacion = convert_to_local_time(orden.get('fecha_anulacion_completa', orden.get('fecha_creacion')))
                fecha_str = fecha_anulacion.strftime('%d/%m/%Y %H:%M') if fecha_anulacion else 'N/A'
                
                ws_anulados.cell(row=row, column=1, value=orden.get('numero_orden', 'N/A'))
                ws_anulados.cell(row=row, column=2, value=mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A'))
                ws_anulados.cell(row=row, column=3, value=meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A'))
                ws_anulados.cell(row=row, column=4, value=item.get('nombre', 'N/A'))
                ws_anulados.cell(row=row, column=5, value=item.get('cantidad', 0))
                ws_anulados.cell(row=row, column=6, value=item.get('precio_unitario', 0))
                ws_anulados.cell(row=row, column=7, value=item.get('cantidad', 0) * item.get('precio_unitario', 0))
                ws_anulados.cell(row=row, column=8, value=fecha_str)
                ws_anulados.cell(row=row, column=9, value=item.get('motivo_anulacion', 'Anulado por orden completa'))
                row += 1
    
    # Ajustar ancho de todas las columnas en todas las hojas
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generar_reporte_excel(ordenes, mesas, meseros, tickets, periodo_nombre, start_date=None, end_date=None):
    """Genera un reporte en Excel con dos hojas: productos y tickets"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # --- Hoja de Productos ---
    ws_productos = wb.active
    ws_productos.title = "productos"
    
    # Encabezados productos
    headers_productos = ["Plato Vendido", "Cantidad", "Precio Venta", "Total Venta", "Fecha de Venta", "Mesero", "Método de Pago"]
    for col, header in enumerate(headers_productos, 1):
        cell = ws_productos.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Datos productos
    row = 2
    for orden in ordenes:
        fecha_venta = convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion')))
        fecha_str = fecha_venta.strftime('%d/%m/%Y %H:%M') if fecha_venta else 'N/A'
        mesero_nombre = meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')
        
        # Buscar el ticket correspondiente para obtener método de pago
        ticket_orden = next((ticket for ticket in tickets if ticket.get('orden_id') == orden.get('_id')), {})
        pago_info = ticket_orden.get('pago_info', {})
        metodo_pago = pago_info.get('metodo', 'No especificado')
        
        # Formatear método de pago para mejor legibilidad
        if metodo_pago == 'efectivo':
            metodo_display = 'Efectivo'
        elif metodo_pago == 'tarjeta':
            tipo_tarjeta = pago_info.get('tipo_tarjeta', '').title()
            metodo_display = f'Tarjeta ({tipo_tarjeta})'
        elif metodo_pago == 'mixto':
            metodo_display = 'Pago Mixto'
        elif metodo_pago == 'transferencia':
            metodo_display = 'Transferencia'
        elif metodo_pago == 'criptomoneda':
            tipo_cripto = pago_info.get('tipo_cripto', 'Criptomoneda')
            metodo_display = f'Cripto ({tipo_cripto})'
        else:
            metodo_display = metodo_pago.title()
        
        for item in orden.get('items', []):
            # Skip cancelled/anulled items
            if item.get('anulado', False):
                continue
            ws_productos.cell(row=row, column=1, value=item.get('nombre', 'N/A'))
            ws_productos.cell(row=row, column=2, value=item.get('cantidad', 0))
            ws_productos.cell(row=row, column=3, value=item.get('precio_unitario', 0))
            ws_productos.cell(row=row, column=4, value=item.get('cantidad', 0) * item.get('precio_unitario', 0))
            ws_productos.cell(row=row, column=5, value=fecha_str)
            ws_productos.cell(row=row, column=6, value=mesero_nombre)
            ws_productos.cell(row=row, column=7, value=metodo_display)
            row += 1
    
    # Ajustar ancho de columnas
    for column in ws_productos.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_productos.column_dimensions[column_letter].width = adjusted_width
    
    # --- Hoja de Tickets ---
    ws_tickets = wb.create_sheet("tickets")
    
    # Encabezados tickets
    headers_tickets = ["Numero Orden", "Numero Ticket", "Mesa", "Mesero", "Total Ticket", "Fecha Venta", "Método de Pago"]
    for col, header in enumerate(headers_tickets, 1):
        cell = ws_tickets.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Datos tickets
    row = 2
    for orden in ordenes:
        fecha_venta = convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion')))
        fecha_str = fecha_venta.strftime('%d/%m/%Y %H:%M') if fecha_venta else 'N/A'
        
        # Buscar el ticket correspondiente para obtener método de pago
        ticket_orden = next((ticket for ticket in tickets if ticket.get('orden_id') == orden.get('_id')), {})
        pago_info = ticket_orden.get('pago_info', {})
        metodo_pago = pago_info.get('metodo', 'No especificado')
        
        # Formatear método de pago para mejor legibilidad
        if metodo_pago == 'efectivo':
            metodo_display = 'Efectivo'
        elif metodo_pago == 'tarjeta':
            tipo_tarjeta = pago_info.get('tipo_tarjeta', '').title()
            metodo_display = f'Tarjeta ({tipo_tarjeta})'
        elif metodo_pago == 'mixto':
            metodo_display = 'Pago Mixto'
        elif metodo_pago == 'transferencia':
            metodo_display = 'Transferencia'
        elif metodo_pago == 'criptomoneda':
            tipo_cripto = pago_info.get('tipo_cripto', 'Criptomoneda')
            metodo_display = f'Cripto ({tipo_cripto})'
        else:
            metodo_display = metodo_pago.title()
        
        ws_tickets.cell(row=row, column=1, value=orden.get('numero_orden', 'N/A'))
        ws_tickets.cell(row=row, column=2, value=orden.get('numero_orden', 'N/A'))  # Usando mismo numero como ticket
        ws_tickets.cell(row=row, column=3, value=mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A'))
        ws_tickets.cell(row=row, column=4, value=meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A'))
        ws_tickets.cell(row=row, column=5, value=orden.get('total', 0))
        ws_tickets.cell(row=row, column=6, value=fecha_str)
        ws_tickets.cell(row=row, column=7, value=metodo_display)
        row += 1
    
    # Ajustar ancho de columnas tickets
    for column in ws_tickets.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_tickets.column_dimensions[column_letter].width = adjusted_width
    
    # Agregar logo si existe y OpenpyxlImage está disponible
    try:
        if OpenpyxlImage is not None:
            logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'LOGO.png')
            if os.path.exists(logo_path):
                img = OpenpyxlImage(logo_path)
                img.height = 80
                img.width = 120
                ws_productos.add_image(img, 'H1')
                ws_tickets.add_image(img, 'H1')
    except:
        pass
    
    # Guardar en memoria
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def limpiar_texto_pdf(texto):
    """Limpia emojis y caracteres no compatibles con PDF"""
    if not texto:
        return "N/A"
    
    import re
    # Remover emojis y caracteres especiales
    texto_limpio = re.sub(r'[^\x00-\x7F]+', '', str(texto))  # Solo ASCII
    # Limpiar espacios extra y caracteres especiales comunes
    texto_limpio = re.sub(r'[🎉🍹🍽️💰🔥⭐💥⏰👨‍🍳✅❌]', '', texto_limpio)
    # Limpiar caracteres de formato especial
    texto_limpio = texto_limpio.strip()
    
    return texto_limpio if texto_limpio else "Producto"

def generar_reporte_pdf_completo(ordenes_pagadas, ordenes_anuladas, mesas, meseros, tickets, periodo_nombre, start_date=None, end_date=None, resumen_stats=None):
    """Genera un reporte completo en PDF con órdenes pagadas, anuladas y estadísticas"""
    
    class ReportePDFCompleto(FPDF):
        def header(self):
            # Logo
            try:
                logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'LOGO.png')
                if os.path.exists(logo_path):
                    self.image(logo_path, 10, 8, 33)
            except:
                pass
            
            self.set_font('Helvetica', 'B', 15)
            self.cell(0, 10, f'Reporte Completo de Ventas - {periodo_nombre}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            if start_date and end_date:
                self.set_font('Helvetica', '', 10)
                periodo_str = f"Del {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
                self.cell(0, 10, periodo_str, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            self.ln(10)
        
        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 10, f'Página {self.page_no()}', align='C')
    
    pdf = ReportePDFCompleto()
    pdf.add_page()
    
    # Resumen general con estadísticas completas
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, 'RESUMEN GENERAL', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(5)
    
    if resumen_stats:
        pdf.set_font('Helvetica', 'B', 11)
        pdf.cell(0, 8, 'VENTAS REALIZADAS', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 6, f'• Órdenes Pagadas: {resumen_stats["ordenes_pagadas"]}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 6, f'• Total Ventas Brutas: ${resumen_stats["ventas_brutas"]:.2f}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 6, f'• Items Vendidos: {resumen_stats["items_vendidos"]}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.ln(3)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.set_text_color(220, 53, 69)  # Color rojo
        pdf.cell(0, 8, 'VENTAS ANULADAS', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(0, 6, f'• Órdenes Anuladas: {resumen_stats["ordenes_anuladas"]}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 6, f'• Total Perdido: ${resumen_stats["total_perdido"]:.2f}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 6, f'• Items Anulados: {resumen_stats["items_anulados"]}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.ln(3)
        pdf.set_font('Helvetica', 'B', 11)
        pdf.set_text_color(56, 142, 60)  # Color verde
        pdf.cell(0, 8, 'RESULTADO NETO', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(0, 6, f'• Ventas Netas: ${resumen_stats["ventas_netas"]:.2f}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.set_text_color(0, 0, 0)  # Volver a negro
    
    pdf.ln(10)
    
    # Sección de Productos Vendidos
    if ordenes_pagadas:
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 10, 'PRODUCTOS VENDIDOS', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
        pdf.ln(5)
        
        # Encabezados productos
        pdf.set_font('Helvetica', 'B', 8)
        pdf.cell(35, 8, 'Producto', border=1, align='C')
        pdf.cell(12, 8, 'Cant.', border=1, align='C')
        pdf.cell(15, 8, 'Precio', border=1, align='C')
        pdf.cell(18, 8, 'Subtotal', border=1, align='C')
        pdf.cell(25, 8, 'Fecha', border=1, align='C')
        pdf.cell(30, 8, 'Mesero', border=1, align='C')
        pdf.cell(35, 8, 'Método Pago', border=1, align='C')
        pdf.ln()
        
        # Datos productos
        pdf.set_font('Helvetica', '', 7)
        for orden in ordenes_pagadas:
            fecha_venta = convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion')))
            fecha_str = fecha_venta.strftime('%d/%m/%Y') if fecha_venta else 'N/A'
            mesero_nombre = meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')[:15]
            
            ticket_orden = next((ticket for ticket in tickets if ticket.get('orden_id') == orden.get('_id')), {})
            metodo_pago = ticket_orden.get('pago_info', {}).get('metodo', 'N/E')
            
            for item in orden.get('items', []):
                if item.get('anulado', False):
                    continue
                
                subtotal = item.get('cantidad', 0) * item.get('precio_unitario', 0)
                
                pdf.cell(35, 6, (item.get('nombre', 'N/A')[:20])[:20], border=1)
                pdf.cell(12, 6, str(item.get('cantidad', 0)), border=1, align='C')
                pdf.cell(15, 6, f"${item.get('precio_unitario', 0):.1f}", border=1, align='R')
                pdf.cell(18, 6, f"${subtotal:.2f}", border=1, align='R')
                pdf.cell(25, 6, fecha_str, border=1, align='C')
                pdf.cell(30, 6, mesero_nombre, border=1)
                pdf.cell(35, 6, metodo_pago.title(), border=1)
                pdf.ln()
        
        pdf.ln(5)
    
    # Sección de Órdenes Anuladas
    if ordenes_anuladas:
        pdf.add_page()
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_text_color(220, 53, 69)  # Color rojo
        pdf.cell(0, 10, 'ÓRDENES ANULADAS', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
        pdf.set_text_color(0, 0, 0)  # Volver a negro
        pdf.ln(5)
        
        # Encabezados anuladas
        pdf.set_font('Helvetica', 'B', 8)
        pdf.cell(20, 8, 'Orden #', border=1, align='C')
        pdf.cell(25, 8, 'Mesa', border=1, align='C')
        pdf.cell(30, 8, 'Mesero', border=1, align='C')
        pdf.cell(20, 8, 'Total', border=1, align='C')
        pdf.cell(25, 8, 'F. Anulación', border=1, align='C')
        pdf.cell(30, 8, 'Anulada Por', border=1, align='C')
        pdf.cell(40, 8, 'Motivo', border=1, align='C')
        pdf.ln()
        
        # Datos anuladas
        pdf.set_font('Helvetica', '', 7)
        for orden in ordenes_anuladas:
            fecha_anulacion = convert_to_local_time(orden.get('fecha_anulacion_completa'))
            fecha_str = fecha_anulacion.strftime('%d/%m/%Y') if fecha_anulacion else 'N/A'
            
            items_no_anulados = [item for item in orden.get('items', []) if not item.get('anulado', False)]
            total_perdido = sum(item.get('cantidad', 0) * item.get('precio_unitario', 0) for item in items_no_anulados)
            
            motivo = orden.get('motivo_solicitud_anulacion_completa', 'N/E')[:25]
            
            pdf.cell(20, 6, str(orden.get('numero_orden', 'N/A')), border=1, align='C')
            pdf.cell(25, 6, mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A')[:15], border=1)
            pdf.cell(30, 6, meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')[:20], border=1)
            pdf.cell(20, 6, f"${total_perdido:.2f}", border=1, align='R')
            pdf.cell(25, 6, fecha_str, border=1, align='C')
            pdf.cell(30, 6, orden.get('usuario_anula_completa', 'N/A')[:20], border=1)
            pdf.cell(40, 6, motivo, border=1)
            pdf.ln()
    
    # Guardar en buffer
    output = io.BytesIO()
    pdf.output(dest='S').encode('latin-1')
    pdf_data = pdf.output(dest='S')
    output.write(pdf_data.encode('latin-1') if isinstance(pdf_data, str) else pdf_data)
    output.seek(0)
    return output.getvalue()

def generar_reporte_pdf(ordenes, mesas, meseros, tickets, periodo_nombre, start_date=None, end_date=None):
    """Genera un reporte en PDF con productos y tickets"""
    
    class ReportePDF(FPDF):
        def header(self):
            # Logo
            try:
                logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'LOGO.png')
                if os.path.exists(logo_path):
                    self.image(logo_path, 10, 8, 33)
            except:
                pass
            
            self.set_font('Helvetica', 'B', 15)
            self.cell(0, 10, f'Reporte de Ventas - {periodo_nombre}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            if start_date and end_date:
                self.set_font('Helvetica', '', 10)
                self.cell(0, 10, f'Del {start_date} al {end_date}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            self.ln(10)
        
        def footer(self):
            self.set_y(-15)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 10, f'Página {self.page_no()}', align='C')
    
    pdf = ReportePDF()
    pdf.add_page()
    
    # Resumen general
    total_ordenes = len(ordenes)
    total_ventas = sum(orden.get('total', 0) for orden in ordenes)
    total_items = sum(sum(item.get('cantidad', 0) for item in orden.get('items', []) if not item.get('anulado', False)) for orden in ordenes)
    
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 10, 'RESUMEN GENERAL', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 8, f'Total de Órdenes: {total_ordenes}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 8, f'Total de Ventas: ${total_ventas:.2f}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 8, f'Total de Items Vendidos: {total_items}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(10)
    
    # Sección de Productos
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 10, 'PRODUCTOS VENDIDOS', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(5)
    
    # Encabezados productos
    pdf.set_font('Helvetica', 'B', 8)
    pdf.cell(30, 8, 'Plato', border=1, align='C')
    pdf.cell(12, 8, 'Cant.', border=1, align='C')
    pdf.cell(18, 8, 'Precio', border=1, align='C')
    pdf.cell(18, 8, 'Total', border=1, align='C')
    pdf.cell(25, 8, 'Fecha', border=1, align='C')
    pdf.cell(30, 8, 'Mesero', border=1, align='C')
    pdf.cell(27, 8, 'Método Pago', border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    
    # Datos productos
    pdf.set_font('Helvetica', '', 7)
    for orden in ordenes:
        fecha_venta = convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion')))
        fecha_str = fecha_venta.strftime('%d/%m/%Y') if fecha_venta else 'N/A'
        mesero_nombre = limpiar_texto_pdf(meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A'))[:15]
        
        # Buscar el ticket correspondiente para obtener método de pago
        ticket_orden = next((ticket for ticket in tickets if ticket.get('orden_id') == orden.get('_id')), {})
        pago_info = ticket_orden.get('pago_info', {})
        metodo_pago = pago_info.get('metodo', 'N/A')
        
        # Formatear método de pago para PDF (más corto)
        if metodo_pago == 'efectivo':
            metodo_display = 'Efectivo'
        elif metodo_pago == 'tarjeta':
            metodo_display = 'Tarjeta'
        elif metodo_pago == 'mixto':
            metodo_display = 'Mixto'
        elif metodo_pago == 'transferencia':
            metodo_display = 'Transfer.'
        elif metodo_pago == 'criptomoneda':
            metodo_display = 'Cripto'
        else:
            metodo_display = 'N/A'
        
        for item in orden.get('items', []):
            # Skip cancelled/anulled items
            if item.get('anulado', False):
                continue
            if pdf.get_y() > 270:  # Nueva página si es necesario
                pdf.add_page()
                # Repetir encabezados
                pdf.set_font('Helvetica', 'B', 8)
                pdf.cell(30, 8, 'Plato', border=1, align='C')
                pdf.cell(12, 8, 'Cant.', border=1, align='C')
                pdf.cell(18, 8, 'Precio', border=1, align='C')
                pdf.cell(18, 8, 'Total', border=1, align='C')
                pdf.cell(25, 8, 'Fecha', border=1, align='C')
                pdf.cell(30, 8, 'Mesero', border=1, align='C')
                pdf.cell(27, 8, 'Método Pago', border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
                pdf.set_font('Helvetica', '', 7)
            
            nombre_item = limpiar_texto_pdf(item.get('nombre', 'N/A'))[:20]
            cantidad = item.get('cantidad', 0)
            precio = item.get('precio_unitario', 0)
            total_item = cantidad * precio
            
            pdf.cell(30, 8, nombre_item, border=1)
            pdf.cell(12, 8, str(cantidad), border=1, align='C')
            pdf.cell(18, 8, f'${precio:.2f}', border=1, align='R')
            pdf.cell(18, 8, f'${total_item:.2f}', border=1, align='R')
            pdf.cell(25, 8, fecha_str, border=1, align='C')
            pdf.cell(30, 8, mesero_nombre, border=1)
            pdf.cell(27, 8, metodo_display, border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    
    # Nueva página para tickets
    pdf.add_page()
    
    # Sección de Tickets
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 10, 'TICKETS/ÓRDENES', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    pdf.ln(5)
    
    # Encabezados tickets
    pdf.set_font('Helvetica', 'B', 8)
    pdf.cell(20, 8, 'Orden #', border=1, align='C')
    pdf.cell(20, 8, 'Ticket #', border=1, align='C')
    pdf.cell(25, 8, 'Mesa', border=1, align='C')
    pdf.cell(30, 8, 'Mesero', border=1, align='C')
    pdf.cell(20, 8, 'Total', border=1, align='C')
    pdf.cell(25, 8, 'Fecha', border=1, align='C')
    pdf.cell(30, 8, 'Método Pago', border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    
    # Datos tickets
    pdf.set_font('Helvetica', '', 8)
    for orden in ordenes:
        if pdf.get_y() > 270:  # Nueva página si es necesario
            pdf.add_page()
            # Repetir encabezados
            pdf.set_font('Helvetica', 'B', 8)
            pdf.cell(20, 8, 'Orden #', border=1, align='C')
            pdf.cell(20, 8, 'Ticket #', border=1, align='C')
            pdf.cell(25, 8, 'Mesa', border=1, align='C')
            pdf.cell(30, 8, 'Mesero', border=1, align='C')
            pdf.cell(20, 8, 'Total', border=1, align='C')
            pdf.cell(25, 8, 'Fecha', border=1, align='C')
            pdf.cell(30, 8, 'Método Pago', border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.set_font('Helvetica', '', 8)
        
        fecha_venta = convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion')))
        fecha_str = fecha_venta.strftime('%d/%m/%Y') if fecha_venta else 'N/A'
        
        # Buscar el ticket correspondiente para obtener método de pago
        ticket_orden = next((ticket for ticket in tickets if ticket.get('orden_id') == orden.get('_id')), {})
        pago_info = ticket_orden.get('pago_info', {})
        metodo_pago = pago_info.get('metodo', 'N/A')
        
        # Formatear método de pago para PDF (más corto)
        if metodo_pago == 'efectivo':
            metodo_display = 'Efectivo'
        elif metodo_pago == 'tarjeta':
            metodo_display = 'Tarjeta'
        elif metodo_pago == 'mixto':
            metodo_display = 'Mixto'
        elif metodo_pago == 'transferencia':
            metodo_display = 'Transfer.'
        elif metodo_pago == 'criptomoneda':
            metodo_display = 'Cripto'
        else:
            metodo_display = 'N/A'
        
        pdf.cell(20, 8, str(orden.get('numero_orden', 'N/A')), border=1, align='C')
        pdf.cell(20, 8, str(orden.get('numero_orden', 'N/A')), border=1, align='C')
        pdf.cell(25, 8, limpiar_texto_pdf(mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A'))[:12], border=1)
        pdf.cell(30, 8, limpiar_texto_pdf(meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A'))[:18], border=1)
        pdf.cell(20, 8, f'${orden.get("total", 0):.2f}', border=1, align='R')
        pdf.cell(25, 8, fecha_str, border=1, align='C')
        pdf.cell(30, 8, metodo_display, border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    
    return bytes(pdf.output())

# --- Estilos CSS ---
st.markdown("""
<style>
    .order-card {
        border: 1px solid #ddd;
        border-radius: 10px;
        padding: 10px;
        margin-bottom: 10px;
        background-color: #f9f9f9;
    }
    .item-row {
        border-bottom: 1px solid #eee;
        padding: 5px 0;
        font-size: 0.9em;
    }
    .summary-card {
        border: 1px solid #ddd;
        border-radius: 10px;
        padding: 15px;
        background-color: #e6f7ff; /* Light blue background for summary */
        margin-bottom: 10px;
        box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
    }
    .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
        font-size:1.2rem;
    }
</style>
""", unsafe_allow_html=True)

# --- Lógica Principal de la Pantalla ---
if 'usuario' in st.session_state:
    db = couchdb_utils.get_database_instance()
    
    if db:
        st.title("📊 Reportes de Ventas")

        # --- Obtener datos base ---
        all_orders = couchdb_utils.get_documents_by_partition(db, "ordenes")
        all_tickets = couchdb_utils.get_documents_by_partition(db, "tickets")
        mesas = {m['_id']: m for m in couchdb_utils.get_documents_by_partition(db, "mesas")}
        meseros = {m['_id']: m for m in couchdb_utils.get_documents_by_partition(db, "Usuario") if m.get('id_rol') == 3}

        # --- Layout Principal: Columnas para Pestañas y Resumen ---
        col_tabs, col_summary = st.columns([2, 1])

        with col_tabs:
            tab_daily_sales, tab_weekly_sales, tab_monthly_sales, tab_in_process_sales, tab_cancelled_orders = st.tabs(["Ventas del Día", "Ventas por Semana", "Ventas por Mes", "Ventas en Proceso", "Ventas Anuladas"])

            with tab_daily_sales:
                st.subheader("Ventas del Día (Órdenes Pagadas)")
                # Filtrar órdenes pagadas del día actual usando zona horaria local
                local_tz = get_timezone()
                today = datetime.now(local_tz).date()
                ordenes_pagadas_hoy = []
                for orden in all_orders:
                    if orden.get('estado') == 'pagada':
                        # Use 'fecha_pago' if available, otherwise 'fecha_creacion'
                        date_to_check_str = orden.get('fecha_pago', orden.get('fecha_creacion'))
                        try:
                            # Convert to local time to get the correct date
                            local_dt = convert_to_local_time(date_to_check_str)
                            if local_dt:
                                order_date = local_dt.date()
                                if order_date == today:
                                    ordenes_pagadas_hoy.append(orden)
                        except:
                            # Handle cases where date conversion fails
                            continue
                
                ordenes_pagadas_hoy.sort(key=lambda x: x.get('numero_orden', 0), reverse=True)

                if not ordenes_pagadas_hoy:
                    st.info("No hay órdenes pagadas para el día de hoy.")
                else:
                    # Botones de descarga
                    col_btn1, col_btn2 = st.columns(2)
                    with col_btn1:
                        if st.button("📊 Descargar Excel - Día", type="secondary", use_container_width=True):
                            excel_data = generar_reporte_excel(
                                ordenes_pagadas_hoy, mesas, meseros, all_tickets,
                                "Ventas del Día", 
                                today.strftime('%d/%m/%Y'), today.strftime('%d/%m/%Y')
                            )
                            st.download_button(
                                label="Descargar Reporte Excel",
                                data=excel_data,
                                file_name=f"Ventas_Dia_{today.strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    with col_btn2:
                        if st.button("📄 Descargar PDF - Día", type="secondary", use_container_width=True):
                            pdf_data = generar_reporte_pdf(
                                ordenes_pagadas_hoy, mesas, meseros, all_tickets,
                                "Ventas del Día",
                                today.strftime('%d/%m/%Y'), today.strftime('%d/%m/%Y')
                            )
                            st.download_button(
                                label="Descargar Reporte PDF",
                                data=pdf_data,
                                file_name=f"Ventas_Dia_{today.strftime('%Y%m%d')}.pdf",
                                mime="application/pdf"
                            )
                    
                    st.markdown("---")
                    for orden in ordenes_pagadas_hoy:
                        # Buscar método de pago para mostrar en el título
                        ticket_orden = next((ticket for ticket in all_tickets if ticket.get('orden_id') == orden.get('_id')), {})
                        pago_info = ticket_orden.get('pago_info', {})
                        metodo_pago = pago_info.get('metodo', 'No especificado')
                        
                        # Formatear método de pago para mostrar
                        if metodo_pago == 'efectivo':
                            metodo_display = 'Efectivo'
                        elif metodo_pago == 'tarjeta':
                            tipo_tarjeta = pago_info.get('tipo_tarjeta', '').title()
                            metodo_display = f'Tarjeta ({tipo_tarjeta})'
                        elif metodo_pago == 'mixto':
                            metodo_display = 'Pago Mixto'
                        elif metodo_pago == 'transferencia':
                            metodo_display = 'Transferencia'
                        elif metodo_pago == 'criptomoneda':
                            tipo_cripto = pago_info.get('tipo_cripto', 'Criptomoneda')
                            metodo_display = f'Cripto ({tipo_cripto})'
                        else:
                            metodo_display = metodo_pago.title()
                        
                        with st.expander(f"Orden #{orden.get('numero_orden', 'N/A')} - Mesa: {mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A')} - Total: ${orden.get('total', 0):.2f} - {metodo_display}", expanded=False):
                            st.markdown(f"""
                            **Mesero:** {meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')}  
                            **Hora Pago:** {convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion'))).strftime('%d/%m/%Y %I:%M %p') if convert_to_local_time(orden.get('fecha_pago', orden.get('fecha_creacion'))) else 'N/A'}  
                            **Método de Pago:** {metodo_display}
                            """)
                            st.markdown("**Productos:**")
                            for item in orden.get('items', []):
                                # Skip cancelled/anulled items
                                if item.get('anulado', False):
                                    continue
                                st.markdown(f"""
                                <div class="item-row">
                                    <b>{item.get('cantidad', 0)}x {item.get('nombre', 'N/A')}</b> (${item.get('precio_unitario', 0):.2f} c/u)
                                    {f"<i>({item.get('comentarios', '')})</i>" if item.get('comentarios') else ""}
                                </div>
                                """, unsafe_allow_html=True)
                            if orden.get('comentarios'):
                                st.markdown(f"**Comentarios de la orden:** {orden['comentarios']}")
                            if orden.get('efectivo_recibido') is not None:
                                st.markdown(f"**Efectivo Recibido:** ${orden['efectivo_recibido']:.2f} | **Cambio:** ${orden['cambio']:.2f}")

            with tab_weekly_sales:
                st.subheader("Ventas de la Semana (Órdenes Pagadas)")
                # Calcular inicio y fin de la semana actual (lunes a domingo) usando zona horaria local
                local_tz = get_timezone()
                today = datetime.now(local_tz).date()
                # Encontrar el lunes de esta semana
                days_since_monday = today.weekday()
                start_of_week = today - timedelta(days=days_since_monday)
                end_of_week = start_of_week + timedelta(days=6)
                
                st.info(f"Semana del {start_of_week.strftime('%d/%m/%Y')} al {end_of_week.strftime('%d/%m/%Y')}")
                
                # Filtrar órdenes pagadas de la semana
                ordenes_semana = []
                for orden in all_orders:
                    if orden.get('estado') == 'pagada':
                        date_to_check_str = orden.get('fecha_pago', orden.get('fecha_creacion'))
                        try:
                            # Convert to local time to get the correct date
                            local_dt = convert_to_local_time(date_to_check_str)
                            if local_dt:
                                order_date = local_dt.date()
                                if start_of_week <= order_date <= end_of_week:
                                    ordenes_semana.append(orden)
                        except:
                            continue
                
                ordenes_semana.sort(key=lambda x: x.get('numero_orden', 0), reverse=True)
                
                if not ordenes_semana:
                    st.info("No hay órdenes pagadas para esta semana.")
                else:
                    # Botones de descarga
                    col_btn1, col_btn2 = st.columns(2)
                    with col_btn1:
                        if st.button("📊 Descargar Excel - Semana", type="secondary", use_container_width=True):
                            excel_data = generar_reporte_excel(
                                ordenes_semana, mesas, meseros, all_tickets,
                                "Ventas de la Semana",
                                start_of_week.strftime('%d/%m/%Y'), end_of_week.strftime('%d/%m/%Y')
                            )
                            st.download_button(
                                label="Descargar Reporte Excel",
                                data=excel_data,
                                file_name=f"Ventas_Semana_{start_of_week.strftime('%Y%m%d')}_{end_of_week.strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    with col_btn2:
                        if st.button("📄 Descargar PDF - Semana", type="secondary", use_container_width=True):
                            pdf_data = generar_reporte_pdf(
                                ordenes_semana, mesas, meseros, all_tickets,
                                "Ventas de la Semana",
                                start_of_week.strftime('%d/%m/%Y'), end_of_week.strftime('%d/%m/%Y')
                            )
                            st.download_button(
                                label="Descargar Reporte PDF",
                                data=pdf_data,
                                file_name=f"Ventas_Semana_{start_of_week.strftime('%Y%m%d')}_{end_of_week.strftime('%Y%m%d')}.pdf",
                                mime="application/pdf"
                            )
                    
                    st.markdown("---")
                    # Resumen de la semana
                    total_semana = sum(orden.get('total', 0) for orden in ordenes_semana)
                    st.metric("Total Ventas de la Semana", f"${total_semana:.2f}")
                    
                    # Agrupar por día
                    ventas_por_dia = {}
                    for orden in ordenes_semana:
                        date_to_check_str = orden.get('fecha_pago', orden.get('fecha_creacion'))
                        try:
                            # Convert to local time to get the correct date
                            local_dt = convert_to_local_time(date_to_check_str)
                            if local_dt:
                                order_date = local_dt.date()
                                dia_nombre = order_date.strftime('%A %d/%m')
                                if dia_nombre not in ventas_por_dia:
                                    ventas_por_dia[dia_nombre] = []
                                ventas_por_dia[dia_nombre].append(orden)
                        except:
                            continue
                    
                    # Mostrar por día
                    for dia, ordenes_dia in sorted(ventas_por_dia.items()):
                        total_dia = sum(orden.get('total', 0) for orden in ordenes_dia)
                        with st.expander(f"{dia} - {len(ordenes_dia)} órdenes - Total: ${total_dia:.2f}", expanded=False):
                            for orden in ordenes_dia:
                                st.markdown(f"**Orden #{orden.get('numero_orden', 'N/A')}** - Mesa: {mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A')} - Mesero: {meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')} - ${orden.get('total', 0):.2f}")

            with tab_monthly_sales:
                st.subheader("Ventas del Mes (Órdenes Pagadas)")
                # Calcular inicio y fin del mes actual usando zona horaria local
                local_tz = get_timezone()
                today = datetime.now(local_tz).date()
                start_of_month = today.replace(day=1)
                # Obtener el último día del mes
                if today.month == 12:
                    end_of_month = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    end_of_month = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
                
                st.info(f"Mes: {start_of_month.strftime('%d/%m/%Y')} al {end_of_month.strftime('%d/%m/%Y')}")
                
                # Filtrar órdenes pagadas del mes
                ordenes_mes = []
                for orden in all_orders:
                    if orden.get('estado') == 'pagada':
                        date_to_check_str = orden.get('fecha_pago', orden.get('fecha_creacion'))
                        try:
                            # Convert to local time to get the correct date
                            local_dt = convert_to_local_time(date_to_check_str)
                            if local_dt:
                                order_date = local_dt.date()
                                if start_of_month <= order_date <= end_of_month:
                                    ordenes_mes.append(orden)
                        except:
                            continue
                
                ordenes_mes.sort(key=lambda x: x.get('numero_orden', 0), reverse=True)
                
                if not ordenes_mes:
                    st.info("No hay órdenes pagadas para este mes.")
                else:
                    # Botones de descarga
                    col_btn1, col_btn2 = st.columns(2)
                    with col_btn1:
                        if st.button("📊 Descargar Excel - Mes", type="secondary", use_container_width=True):
                            excel_data = generar_reporte_excel(
                                ordenes_mes, mesas, meseros, all_tickets,
                                "Ventas del Mes",
                                start_of_month.strftime('%d/%m/%Y'), end_of_month.strftime('%d/%m/%Y')
                            )
                            st.download_button(
                                label="Descargar Reporte Excel",
                                data=excel_data,
                                file_name=f"Ventas_Mes_{start_of_month.strftime('%Y%m')}_{end_of_month.strftime('%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    
                    with col_btn2:
                        if st.button("📄 Descargar PDF - Mes", type="secondary", use_container_width=True):
                            pdf_data = generar_reporte_pdf(
                                ordenes_mes, mesas, meseros, all_tickets,
                                "Ventas del Mes",
                                start_of_month.strftime('%d/%m/%Y'), end_of_month.strftime('%d/%m/%Y')
                            )
                            st.download_button(
                                label="Descargar Reporte PDF",
                                data=pdf_data,
                                file_name=f"Ventas_Mes_{start_of_month.strftime('%Y%m')}_{end_of_month.strftime('%d')}.pdf",
                                mime="application/pdf"
                            )
                    
                    st.markdown("---")
                    # Resumen del mes
                    total_mes = sum(orden.get('total', 0) for orden in ordenes_mes)
                    st.metric("Total Ventas del Mes", f"${total_mes:.2f}")
                    
                    # Agrupar por día
                    ventas_por_dia_mes = {}
                    for orden in ordenes_mes:
                        date_to_check_str = orden.get('fecha_pago', orden.get('fecha_creacion'))
                        try:
                            # Convert to local time to get the correct date
                            local_dt = convert_to_local_time(date_to_check_str)
                            if local_dt:
                                order_date = local_dt.date()
                                dia_key = order_date.strftime('%Y-%m-%d')
                                dia_display = order_date.strftime('%d/%m/%Y')
                                if dia_key not in ventas_por_dia_mes:
                                    ventas_por_dia_mes[dia_key] = {'display': dia_display, 'ordenes': []}
                                ventas_por_dia_mes[dia_key]['ordenes'].append(orden)
                        except:
                            continue
                    
                    # Mostrar por día (ordenado por fecha)
                    for dia_key in sorted(ventas_por_dia_mes.keys(), reverse=True):
                        dia_data = ventas_por_dia_mes[dia_key]
                        ordenes_dia = dia_data['ordenes']
                        total_dia = sum(orden.get('total', 0) for orden in ordenes_dia)
                        with st.expander(f"{dia_data['display']} - {len(ordenes_dia)} órdenes - Total: ${total_dia:.2f}", expanded=False):
                            for orden in ordenes_dia:
                                st.markdown(f"**Orden #{orden.get('numero_orden', 'N/A')}** - Mesa: {mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A')} - Mesero: {meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')} - ${orden.get('total', 0):.2f}")

            with tab_in_process_sales:
                st.subheader("Ventas en Proceso (Órdenes Pendientes/En Cobro)")
                # Filtrar órdenes pendientes o en cobro
                ordenes_en_proceso = [
                    orden for orden in all_orders 
                    if orden.get('estado') in ['pendiente', 'en_cobro']
                ]
                ordenes_en_proceso.sort(key=lambda x: x.get('numero_orden', 0), reverse=True)

                if not ordenes_en_proceso:
                    st.info("No hay órdenes en proceso actualmente.")
                else:
                    for orden in ordenes_en_proceso:
                        estado_display = "En Proceso (Pendiente)" if orden.get('estado') == 'pendiente' else "En Cobro"
                        with st.expander(f"Orden #{orden.get('numero_orden', 'N/A')} - Mesa: {mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A')} - Estado: {estado_display} - Total: ${orden.get('total', 0):.2f}", expanded=False):
                            st.markdown(f"""
                            **Mesero:** {meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')}  
                            **Hora Creación:** {convert_to_local_time(orden.get('fecha_creacion')).strftime('%d/%m/%Y %I:%M %p') if convert_to_local_time(orden.get('fecha_creacion')) else 'N/A'}
                            """)
                            st.markdown("**Productos:**")
                            for item in orden.get('items', []):
                                # Skip cancelled/anulled items
                                if item.get('anulado', False):
                                    continue
                                st.markdown(f"""
                                <div class="item-row">
                                    <b>{item.get('cantidad', 0)}x {item.get('nombre', 'N/A')}</b> (${item.get('precio_unitario', 0):.2f} c/u)
                                    {f"<i>({item.get('comentarios', '')})</i>" if item.get('comentarios') else ""}
                                </div>
                                """, unsafe_allow_html=True)
                            if orden.get('comentarios'):
                                st.markdown(f"**Comentarios de la orden:** {orden['comentarios']}")

            with tab_cancelled_orders:
                st.subheader("🗑️ Ventas Anuladas (Órdenes Canceladas)")
                
                # Filtrar órdenes anuladas
                ordenes_anuladas = [
                    orden for orden in all_orders 
                    if orden.get('estado') == 'anulada'
                ]
                
                if not ordenes_anuladas:
                    st.info("No hay ventas anuladas registradas.")
                else:
                    # Filtros para las órdenes anuladas
                    col_search, col_date_filter = st.columns([1, 1])
                    
                    with col_search:
                        search_term = st.text_input("🔍 Buscar por número de orden:", placeholder="Ej: 123")
                    
                    with col_date_filter:
                        dias_filtro = st.selectbox("📅 Mostrar anulaciones de:", 
                                                 ["Todas", "Último día", "Últimos 7 días", "Últimos 30 días"])
                    
                    # Aplicar filtros
                    if search_term:
                        ordenes_anuladas = [
                            orden for orden in ordenes_anuladas 
                            if str(search_term).lower() in str(orden.get('numero_orden', '')).lower()
                        ]
                    
                    if dias_filtro != "Todas":
                        local_tz = get_timezone()
                        hoy = datetime.now(local_tz).date()
                        
                        if dias_filtro == "Último día":
                            fecha_limite = hoy - timedelta(days=1)
                        elif dias_filtro == "Últimos 7 días":
                            fecha_limite = hoy - timedelta(days=7)
                        else:  # Últimos 30 días
                            fecha_limite = hoy - timedelta(days=30)
                        
                        ordenes_filtradas = []
                        for orden in ordenes_anuladas:
                            fecha_anulacion_str = orden.get('fecha_anulacion_completa', orden.get('fecha_creacion'))
                            try:
                                local_dt = convert_to_local_time(fecha_anulacion_str)
                                if local_dt and local_dt.date() >= fecha_limite:
                                    ordenes_filtradas.append(orden)
                            except:
                                continue
                        ordenes_anuladas = ordenes_filtradas
                    
                    # Ordenar por número de orden descendente
                    ordenes_anuladas.sort(key=lambda x: x.get('numero_orden', 0), reverse=True)
                    
                    st.success(f"📋 Mostrando {len(ordenes_anuladas)} orden(es) anulada(s)")
                    
                    for orden in ordenes_anuladas:
                        # Obtener información de la anulación
                        fecha_anulacion = convert_to_local_time(orden.get('fecha_anulacion_completa'))
                        fecha_anulacion_str = fecha_anulacion.strftime('%d/%m/%Y %I:%M %p') if fecha_anulacion else 'N/A'
                        usuario_anulo = orden.get('usuario_anula_completa', 'N/A')
                        
                        # Obtener información de la solicitud original
                        motivo_original = orden.get('motivo_solicitud_anulacion_completa', 'No especificado')
                        usuario_solicito = orden.get('usuario_solicita_anulacion_completa', 'N/A')
                        
                        # Calcular total que se perdió
                        items_anulados = [item for item in orden.get('items', []) if not item.get('anulado', False)]
                        total_perdido = sum(item.get('cantidad', 0) * item.get('precio_unitario', 0) for item in items_anulados)
                        
                        with st.expander(f"🗑️ Orden #{orden.get('numero_orden', 'N/A')} - Mesa: {mesas.get(orden.get('mesa_id'), {}).get('descripcion', 'N/A')} - Total Perdido: ${total_perdido:.2f}", expanded=False):
                            col_info1, col_info2 = st.columns(2)
                            
                            with col_info1:
                                st.markdown(f"""
                                **📋 Información de la Orden:**
                                - **Mesero:** {meseros.get(orden.get('mesero_id'), {}).get('nombre', 'N/A')}
                                - **Fecha Creación:** {convert_to_local_time(orden.get('fecha_creacion')).strftime('%d/%m/%Y %I:%M %p') if convert_to_local_time(orden.get('fecha_creacion')) else 'N/A'}
                                - **Total Original:** ${orden.get('total', 0):.2f}
                                """)
                            
                            with col_info2:
                                st.markdown(f"""
                                **🗑️ Información de Anulación:**
                                - **Fecha Anulación:** {fecha_anulacion_str}
                                - **Anulada por:** {usuario_anulo}
                                - **Solicitada por:** {usuario_solicito}
                                """)
                            
                            st.markdown("**💬 Motivo de la Anulación:**")
                            st.info(motivo_original)
                            
                            st.markdown("**🍽️ Productos que se Anularon:**")
                            for item in items_anulados:
                                subtotal = item.get('cantidad', 0) * item.get('precio_unitario', 0)
                                st.markdown(f"""
                                <div class="item-row" style="background-color: #ffebee; border-left: 4px solid #f44336; padding: 8px; margin: 4px 0;">
                                    <b>{item.get('cantidad', 0)}x {item.get('nombre', 'N/A')}</b> - ${item.get('precio_unitario', 0):.2f} c/u = <b>${subtotal:.2f}</b>
                                    {f"<i>({item.get('comentarios', '')})</i>" if item.get('comentarios') else ""}
                                </div>
                                """, unsafe_allow_html=True)
                            
                            # Mostrar información sobre reversión de inventario si está disponible
                            if orden.get('items_revertidos_inventario'):
                                st.markdown("**📦 Inventario Revertido:**")
                                items_revertidos = orden.get('items_revertidos_inventario', [])
                                st.success(f"✅ {len(items_revertidos)} item(s) revertido(s) al inventario")
                                
                                with st.expander("Ver detalles de reversión", expanded=False):
                                    for item_rev in items_revertidos:
                                        st.text(f"• {item_rev.get('item_nombre', 'N/A')}: {item_rev.get('cantidad_revertida', 0)} unidades revertidas")

        with col_summary:
            st.subheader("📊 Resumen de Ventas")
            st.markdown("---")

            st.write("Selecciona el rango de fechas para el resumen:")
            col_date_start, col_date_end = st.columns(2)
            local_tz = get_timezone()
            local_today = datetime.now(local_tz).date()
            with col_date_start:
                start_date = st.date_input("Fecha Inicio", value=local_today - timedelta(days=7))
            with col_date_end:
                end_date = st.date_input("Fecha Fin", value=local_today)

            # Filter paid orders by the selected date range
            ordenes_resumen = []
            for orden in all_orders:
                if orden.get('estado') == 'pagada':
                    date_to_check_str = orden.get('fecha_pago', orden.get('fecha_creacion'))
                    try:
                        # Convert to local time to get the correct date
                        local_dt = convert_to_local_time(date_to_check_str)
                        if local_dt:
                            order_date = local_dt.date()
                            if start_date <= order_date <= end_date:
                                ordenes_resumen.append(orden)
                    except:
                        continue
            
            # Filter cancelled orders by the selected date range
            ordenes_anuladas_resumen = []
            for orden in all_orders:
                if orden.get('estado') == 'anulada':
                    date_to_check_str = orden.get('fecha_anulacion_completa', orden.get('fecha_creacion'))
                    try:
                        # Convert to local time to get the correct date
                        local_dt = convert_to_local_time(date_to_check_str)
                        if local_dt:
                            order_date = local_dt.date()
                            if start_date <= order_date <= end_date:
                                ordenes_anuladas_resumen.append(orden)
                    except:
                        continue
            
            total_ventas_resumen = sum(orden.get('total', 0) for orden in ordenes_resumen)
            total_items_vendidos_resumen = sum(sum(item.get('cantidad', 0) for item in orden.get('items', []) if not item.get('anulado', False)) for orden in ordenes_resumen)
            
            # Calcular totales de órdenes anuladas
            total_anuladas_resumen = len(ordenes_anuladas_resumen)
            total_perdido_anulaciones = sum(
                sum(item.get('cantidad', 0) * item.get('precio_unitario', 0) for item in orden.get('items', []) if not item.get('anulado', False)) 
                for orden in ordenes_anuladas_resumen
            )
            total_items_anulados_resumen = sum(sum(item.get('cantidad', 0) for item in orden.get('items', []) if not item.get('anulado', False)) for orden in ordenes_anuladas_resumen)

            st.markdown(f"""
            <div class="summary-card">
                <p><strong>Período:</strong> {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}</p>
                <hr style="margin: 10px 0; border: 1px solid #ddd;">
                <p><strong>📊 Órdenes Pagadas:</strong> {len(ordenes_resumen)}</p>
                <p><strong>💰 Total de Ventas Brutas:</strong> ${total_ventas_resumen:.2f}</p>
                <p><strong>🍽️ Total de Ítems Vendidos:</strong> {total_items_vendidos_resumen}</p>
                <hr style="margin: 10px 0; border: 1px solid #ddd;">
                <p><strong>🗑️ Órdenes Anuladas:</strong> <span style="color: #d32f2f;">{total_anuladas_resumen}</span></p>
                <p><strong>💸 Total Perdido por Anulaciones:</strong> <span style="color: #d32f2f;">${total_perdido_anulaciones:.2f}</span></p>
                <p><strong>🚫 Ítems Anulados:</strong> <span style="color: #d32f2f;">{total_items_anulados_resumen}</span></p>
                <hr style="margin: 10px 0; border: 1px solid #ddd;">
                <p><strong>📈 Ventas Netas:</strong> <span style="color: #388e3c;">${(total_ventas_resumen - total_perdido_anulaciones):.2f}</span></p>
            </div>
            """, unsafe_allow_html=True)

            if not ordenes_resumen:
                st.info("No hay datos de ventas para generar reportes en el rango seleccionado.")
            else:
                # Botones de descarga para el resumen
                col_btn1, col_btn2 = st.columns(2)
                
                with col_btn1:
                    if st.button("📊 Descargar Excel - Resumen", type="primary", use_container_width=True):
                        # Preparar estadísticas para Excel
                        resumen_stats = {
                            'ordenes_pagadas': len(ordenes_resumen),
                            'ventas_brutas': total_ventas_resumen,
                            'items_vendidos': total_items_vendidos_resumen,
                            'ordenes_anuladas': total_anuladas_resumen,
                            'total_perdido': total_perdido_anulaciones,
                            'items_anulados': total_items_anulados_resumen,
                            'ventas_netas': total_ventas_resumen - total_perdido_anulaciones
                        }
                        
                        excel_data = generar_reporte_excel_completo(
                            ordenes_resumen, ordenes_anuladas_resumen, mesas, meseros, all_tickets,
                            f"Resumen de Ventas",
                            start_date, end_date,
                            resumen_stats
                        )
                        st.download_button(
                            label="Descargar Reporte Excel",
                            data=excel_data,
                            file_name=f"Resumen_Ventas_{start_date.strftime('%Y%m%d')}_a_{end_date.strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                with col_btn2:
                    if st.button("📄 Descargar PDF - Resumen", type="primary", use_container_width=True):
                        # Usar las mismas estadísticas que el Excel
                        resumen_stats = {
                            'ordenes_pagadas': len(ordenes_resumen),
                            'ventas_brutas': total_ventas_resumen,
                            'items_vendidos': total_items_vendidos_resumen,
                            'ordenes_anuladas': total_anuladas_resumen,
                            'total_perdido': total_perdido_anulaciones,
                            'items_anulados': total_items_anulados_resumen,
                            'ventas_netas': total_ventas_resumen - total_perdido_anulaciones
                        }
                        
                        pdf_data = generar_reporte_pdf_completo(
                            ordenes_resumen, ordenes_anuladas_resumen, mesas, meseros, all_tickets,
                            f"Resumen de Ventas",
                            start_date, end_date,
                            resumen_stats
                        )
                        st.download_button(
                            label="Descargar Reporte PDF",
                            data=pdf_data,
                            file_name=f"Resumen_Ventas_{start_date.strftime('%Y%m%d')}_a_{end_date.strftime('%Y%m%d')}.pdf",
                            mime="application/pdf"
                        )
                        
                st.success("Reportes disponibles para descarga.")

    else:
        st.error("No se pudo conectar a la base de datos.")
else:
    st.info("Por favor, inicia sesión para acceder a esta página.")